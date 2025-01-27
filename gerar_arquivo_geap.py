import sys, os
import openpyxl as xl

def main():
    
    ws_arquivo, ws_pgto, wb = inicio()
    ws_arquivo_format, _ = copiarRows(ws_arquivo, ws_pgto)
    
    #debug
    for row in ws_arquivo_format.iter_rows(min_row=1, max_row=20, min_col=1, max_col=5):
        print([cell.value for cell in row])      
    
    #salvar salvar arquivo
    wb.save('geap_locomocao.xlsx')
    
    
def inicio():
    
    """
    Define workbook (file) and worksheet (sheet) variables;
    Clears the cells of the file sheet and then names the columns.
    
    """
    
    #abrir arquivo e carregar planilhas
    wb = xl.load_workbook('geap_locomocao.xlsx', data_only=True)
    ws_pgto = wb['Informação Pagamento']
    ws_arquivo = wb['geap_locomocao']
    
    #limpar celulas para recomeçar
    for row in ws_arquivo.iter_rows():
        for cell in row:
            cell.value = None
    
    #preparar planilha arquivo; nomear colunas
    ws_arquivo['A1'].value = 'MAT'
    ws_arquivo['B1'].value = 'NOME'
    ws_arquivo['C1'].value = 'HRS'
    ws_arquivo['D1'].value = 'TIPO'
    ws_arquivo['E1'].value = 'COMPETENCIA'
    
    return ws_arquivo, ws_pgto, wb
    
    
def copiarRows(ws_arquivo, ws_pgto):
    
    """
    Copies the rows from the user typed worksheet to the file worksheet.
    
    Parameters:
    ws_arquivo: The worksheet where data will be copied to.
    ws_pgto: The worksheet from which data will be copied.
    
    """
    
    #definir ultima matricula cadastrada
    fim_mat_pgto = 20
    
    #buscar os valores da planilha preenchida pelo usuario
    for row in ws_pgto.iter_rows(14, fim_mat_pgto, 1, 5):
        for cell in row:
             ws_arquivo[cell.coordinate].value = cell.value
             
    ws_arquivo.delete_rows(2, 12)
    
    return ws_arquivo, ws_pgto


if __name__ == '__main__':
    main()